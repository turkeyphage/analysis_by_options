import os
import os.path
import platform
import datetime
from datetime import time
import xlrd as xl
import openpyxl as pyxl
from sqlite_manager import SQLiteManager


class FileAnalyzer(SQLiteManager):

    def __init__(self):
        super().__init__()
        # self.thu1 = thulac.thulac(filt=True)
        # self.printFilename = True
        # self.sqlMa = SQLiteManager()
        # self.name_lexical = []
        # self.all_lexical = []
        # self.currFile = ""
        # self.currSheet = ""

        # print(self.all_lexical)
        # for ele in self.sqlMa.execute("SELECT * FROM lexical;"):
        #     # self.all_lexical.append(ele)
        #     for ele_in_tuple in ele:
        #         self.all_lexical.append(ele_in_tuple)

    #查metadata建立日期
    def creation_date(self, path_to_file):
        if platform.system() == 'Windows':
            return os.path.getctime(path_to_file)
        else:
            stat = os.stat(path_to_file)
            try:
                full_datetime = datetime.datetime.fromtimestamp(
                    stat.st_birthtime)
                year = full_datetime.year
                month = full_datetime.month
                day = full_datetime.day
                return "-".join([str(year), str(month), str(day)])
            except AttributeError:
                # We're probably on Linux. No easy way to get creation dates here,
                # so we'll settle for when its content was last modified.
                full_datetime = datetime.datetime.fromtimestamp(stat.st_mtime)
                year = full_datetime.year
                month = full_datetime.month
                day = full_datetime.day
                return "-".join([str(year), str(month), str(day)])

    def parse_excel_file(self, filepath):
        def create_task_table(task_table_name):
                sql_query = "CREATE TABLE IF NOT EXISTS %s (doc_name TEXT, sheet_name TEXT, task_id INT, employee_name TEXT, task_start_date TEXT, task_start_time TEXT, task_end_date TEXT, task_end_time TEXT, task_code TEXT, task_type_name TEXT, task_category_name TEXT, task_desc TEXT, text_full_content TEXT);" % (
                    task_table_name)
                self.execute(sql_query)
                sql_query = "DELETE FROM %s;" % (task_table_name)
                self.execute(sql_query)

        def insert_task_table(task_table_name, valueList):
                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d, '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s');" % (
                    task_table_name, valueList[0], valueList[1], valueList[2], valueList[3], valueList[4], valueList[5], valueList[6], valueList[7], valueList[8], valueList[9], valueList[10], valueList[11], valueList[12])
                self.execute(sql_query)

        filename = os.path.basename(filepath)
        if filename.endswith(".xls") or filename.endswith(".xlsx"):

            # 使用xlrd
            if filename.endswith(".xls"):
                filename_without_extension = filename.replace(
                    ".xls", "").strip()
                task_table_name = "_".join(
                    ["task", filename_without_extension])
                create_task_table(task_table_name)

                excelFile = xl.open_workbook(filepath)
                sheet_names = excelFile.sheet_names()  # 所有sheet的名字

                # sheet iteration
                for i in range(len(sheet_names)):
                    current_sheet = excelFile.sheet_by_index(i)
                    num_rows = current_sheet.nrows
                    num_columns = current_sheet.ncols
                    # 以row數來判斷sheet是否為Empty, 如果empty就不理會
                    if num_rows != 0:
                        curr_sheetname = sheet_names[i].strip()  # sheet名
                        #xlrd的row/column從0開始算
                        for curr_row in range(1, num_rows):

                            doc_name = filename
                            sheet_name = curr_sheetname
                            task_id = curr_row
                            employee_name = ""
                            task_start_date = ""
                            task_start_time = ""
                            task_end_date = ""
                            task_end_time = ""
                            task_code = ""
                            task_type_name = ""
                            task_category_name = ""
                            task_desc = ""
                            text_full_content = ""

                            for curr_column in range(0, num_columns):
                                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                                cell_type = current_sheet.cell_type(
                                    curr_row, curr_column)
                                cell_value = current_sheet.cell_value(
                                    curr_row, curr_column)

                                try:
                                    if cell_type == 3:
                                        cell_value = int(
                                            cell_value * 24 * 3600)
                                        cell_value = time(
                                            cell_value // 3600, (cell_value % 3600) // 60, cell_value % 60)
                                except:
                                    cell_value = ""

                                if curr_column == 0:
                                    task_code = cell_value
                                if curr_column == 2:
                                    employee_name = cell_value
                                if curr_column == 3:
                                    task_start_date = cell_value
                                if curr_column == 4:
                                    task_start_time = cell_value
                                if curr_column == 5:
                                    task_end_date = cell_value
                                if curr_column == 6:
                                    task_end_time = cell_value
                                if curr_column == 9:
                                    task_type_name = cell_value
                                if curr_column == 11:
                                    task_category_name = cell_value.strip().replace(
                                        "'", "").replace("\"", "").upper()
                                if curr_column == 12:
                                    task_desc = cell_value.strip().replace("'", "").replace("\"", "").upper()

                                text_full_content = " ".join([text_full_content, str(cell_value)]).strip(
                                ).replace("'", "").replace("\"", "").upper()

                            insert_task_table(task_table_name, [doc_name, sheet_name, task_id, employee_name, task_start_date, task_start_time,
                                                                task_end_date, task_end_time, task_code, task_type_name, task_category_name, task_desc, text_full_content])

            # 使用openpyxl
            else:
                filename_without_extension = filename.replace(
                    ".xlsx", "").strip()
                task_table_name = "_".join(
                    ["task", filename_without_extension])
                create_task_table(task_table_name)
                excelFile = pyxl.load_workbook(filepath)
                sheet_names = excelFile.get_sheet_names()  # 所有sheet的名字

                #sheet iteration
                for i in range(len(sheet_names)):
                    current_sheet = excelFile.get_sheet_by_name(sheet_names[i])
                    num_rows = current_sheet.max_row
                    num_columns = current_sheet.max_column

                    # 以row跟column數來判斷sheet是否為Empty, 如果empty就不理會
                    if num_columns != 1 and num_rows != 1:
                        curr_sheetname = sheet_names[i].strip()  # sheet名
                        #openpyxl的row/column從1開始算
                        for curr_row in range(2, num_rows + 1):

                            # doc_name TEXT,
                            # sheet_name TEXT,
                            # task_id INT,
                            # employee_name TEXT,
                            # task_start_date TEXT,
                            # task_start_time TEXT,
                            # task_end_date TEXT,
                            # task_end_time TEXT,
                            # task_code TEXT,
                            # task_type_name TEXT,
                            # task_category_name TEXT,
                            # task_desc TEXT,
                            # text_full_content TEXT

                            doc_name = filename
                            sheet_name = curr_sheetname
                            task_id = curr_row
                            employee_name = ""
                            task_start_date = ""
                            task_start_time = ""
                            task_end_date = ""
                            task_end_time = ""
                            task_code = ""
                            task_type_name = ""
                            task_category_name = ""
                            task_desc = ""
                            text_full_content = ""

                            for curr_column in range(1, num_columns + 1):

                                cell_value = current_sheet.cell(
                                    row=curr_row, column=curr_column).value

                                if cell_value == None:
                                    cell_value = ""

                                if curr_column == 1:
                                    task_code = cell_value
                                if curr_column == 3:
                                    employee_name = cell_value
                                if curr_column == 4:
                                    task_start_date = cell_value
                                if curr_column == 5:
                                    task_start_time = cell_value
                                if curr_column == 6:
                                    task_end_date = cell_value
                                if curr_column == 7:
                                    task_end_time = cell_value
                                if curr_column == 10:
                                    task_type_name = cell_value
                                if curr_column == 12:
                                    task_category_name = cell_value.strip().replace(
                                        "'", "").replace("\"", "").upper()
                                if curr_column == 13:
                                    task_desc = cell_value.strip().replace("'", "").replace("\"", "").upper()

                                text_full_content = " ".join([text_full_content, str(cell_value)]).strip(
                                ).replace("'", "").replace("\"", "").upper()

                            insert_task_table(task_table_name, [doc_name, sheet_name, task_id, employee_name, task_start_date, task_start_time,
                                                                task_end_date, task_end_time, task_code, task_type_name, task_category_name, task_desc, text_full_content])

    """
    def count_frequency(self, full_text):

        split_text = full_text.split(" ")

        feq_count = dict()

        for each in split_text:
            if each not in feq_count:
                feq_count[each] = 0

            feq_count[each] = feq_count[each] + 1

        feq_combined = dict()
        for w in sorted(feq_count, key=feq_count.get, reverse=True):
            if feq_count[w] not in feq_combined:
                feq_combined[feq_count[w]] = ""
            feq_combined[feq_count[w]] = ",".join([feq_combined[feq_count[w]], w]).lstrip(',')

        return feq_combined


    #列印結果
    def print_out_result(self, allsheets):
        
        for i in range(len(allsheets)):
            sheet_name = allsheets[i][0]
            od = collections.OrderedDict(sorted(allsheets[i][1].items()))
            # print("sheet_name: " + sheet_name)
            print("<" + sheet_name + ">")
            print("字頻統計: ")
            for k, v in od.items():
                if k > 5:
                    print(str(k) + "次: " + v)
            print("-----------------------------------------------------")



    #單一檔案讀取
    def read_single_file(self, filepath):
        all_sheets = []
        # thu1 = thulac.thulac(seg_only=True, filt=True)
        
        filename = os.path.basename(filepath)

        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            if self.printFilename:
                print("檔案名稱:", filename)
            # thu1 = thulac.thulac(filt=True)

            if filename.endswith(".xls"):
                # 使用xlrd
                
                filename_without_extension = filename.replace(".xls", "").strip()
                # fileDic[filename_without_extension] = []

                excelFile = xl.open_workbook(filepath)
                # 所有sheet的名字
                sheet_names = excelFile.sheet_names()

                # sheet iteration
                for i in range(len(sheet_names)):
                    # sheet
                    work_sheet = excelFile.sheet_by_index(i)

                    # 以row數來判斷sheet是否為Empty
                    if work_sheet.nrows != 0:
                        # print("%s - %s: %d rows" %(filename, sheet_names[i], work_sheet.nrows))
                        # file_create_date = creation_date(src)
                        sheet_fn = sheet_names[i].strip()

                        sheet_contents = ""
                        num_rows = work_sheet.nrows - 1
                        num_cells = work_sheet.ncols - 1
                        curr_row = -1

                        while curr_row < num_rows:
                            curr_row += 1
                            # row = work_sheet.row(curr_row)
                            curr_cell = -1

                            while curr_cell < num_cells:
                                curr_cell += 1
                                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                                cell_type = work_sheet.cell_type(curr_row, curr_cell)
                                cell_value = work_sheet.cell_value(curr_row, curr_cell)

                                try:
                                    if cell_type == 3:
                                        cell_value = int(cell_value * 24 * 3600)
                                        cell_value = time(cell_value // 3600, (cell_value % 3600) // 60, cell_value % 60)
                                except:
                                    cell_value = ""

                                if cell_type != 0 and cell_type != 6 and cell_type != 5:
                                    cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                    sheet_contents = " ".join(
                                        [sheet_contents, str(cell_value)]).strip()

                        # text_cut = thu1.cut(sheet_contents, text=True).strip()

                        text_cut = self.thu1.cut(sheet_contents) #[(tuple)]


                        filted_str = ""
                        for ele in text_cut:
                            if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                filted_str = " ".join([filted_str, ele[0]])
                        text_cut = filted_str.strip()
                        freq_result = self.count_frequency(text_cut)
                        
                        all_sheets.append((sheet_names[i], freq_result))


            else:
                # 使用openpyxl
                filename_without_extension = filename.replace(".xlsx", "").strip()

                excelFile = pyxl.load_workbook(filepath)
                #所有sheet的名字
                sheet_names = excelFile.get_sheet_names()

                #sheet iteration
                for i in range(len(sheet_names)):
                    #sheet
                    work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                    num_rows = work_sheet.max_row
                    num_columns = work_sheet.max_column

                    if num_columns != 1 and num_rows != 1:
                        # not empty
                        sheet_contents = ""

                        for curr_row in range(1, num_rows + 1):
                            for curr_cell in range(1, num_columns + 1):
                                cell_value = work_sheet.cell(
                                    row=curr_row, column=curr_cell).value
                                if cell_value == None:
                                    cell_value = ""

                                cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                sheet_contents = " ".join(
                                    [sheet_contents, str(cell_value)]).strip()
                                
                        # text_cut = thu1.cut(sheet_contents, text=True).strip()
                        text_cut = self.thu1.cut(sheet_contents)
                        filted_str = ""
                        for ele in text_cut:
                            if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                filted_str = " ".join([filted_str, ele[0]])
                        text_cut = filted_str.strip()
                        freq_result = self.count_frequency(text_cut)

                        all_sheets.append((sheet_names[i], freq_result))

        return all_sheets  # [(2 items tuple)]


    """

    """    
    #舊版本
    def parse_row_by_row(self, filepath):
        
        filename = os.path.basename(filepath)

        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            if self.printFilename:
                print("檔案名稱:", filename)

            if filename.endswith(".xls"): # 使用xlrd
                filename_without_extension = filename.replace(".xls", "").strip()
                self.currFile = filename_without_extension

                excelFile = xl.open_workbook(filepath)
                # 所有sheet的名字
                sheet_names = excelFile.sheet_names()

                # sheet iteration
                for i in range(len(sheet_names)):
                    # sheet
                    
                    work_sheet = excelFile.sheet_by_index(i)

                    # 以row數來判斷sheet是否為Empty, 如果empty就不理會
                    if work_sheet.nrows != 0:

                        #sheet名
                        sheet_fn = sheet_names[i].strip()
                        self.currSheet = sheet_fn

                        num_rows = work_sheet.nrows
                        num_columns = work_sheet.ncols
                        
                        lexicon_table_name = "_".join([filename_without_extension, sheet_fn, "lexicon"])
                        thulac_table_name = "_".join([filename_without_extension, sheet_fn, "thulac"])
                        task_table_name = "_".join([filename_without_extension, "task"])

                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)

                        sql_query = "DELETE FROM %s;" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)


                        #create task table
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (doc_name TEXT, sheet_name TEXT, task_id INT, task_full_text TEXT, task_status INT, task_start_date TEXT, task_end_date TEXT, task_code TEXT);" % (task_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (task_table_name)
                        self.sqlMa.execute(sql_query)

                        parsing_result_thulac = dict()
                        parsing_result_lexicon = dict()

                        for curr_row in range(1, num_rows):

                            doc_name = filename
                            sheet_name = sheet_fn
                            task_id = curr_row+1
                            task_full_text = ""
                            task_code = ""
                            task_status = 1
                            task_start_date = ""
                            task_end_date = ""


                            for curr_cell in range(0, num_columns):
                                
                                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                                cell_type = work_sheet.cell_type(curr_row, curr_cell)
                                cell_value = work_sheet.cell_value(curr_row, curr_cell)

                                try:
                                    if cell_type == 3:
                                        cell_value = int(cell_value * 24 * 3600)
                                        cell_value = time(cell_value // 3600, (cell_value % 3600) // 60, cell_value % 60)
                                except:
                                    cell_value = ""

                                task_full_text = " ".join([task_full_text, str(cell_value)]).strip().replace("'", "").replace("\"", "").upper()

                                if curr_cell == 0:
                                    task_code = cell_value

                                if curr_cell == 3:
                                    task_start_date = cell_value
                                
                                if curr_cell == 5:
                                    task_end_date = cell_value




                                if cell_type != 0 and cell_type != 6 and cell_type != 5:
                                    cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value)).upper()
                                    
                                    #for lecial search
                                    for lexicon_tuple in self.all_lexical:
                                        lexicon = lexicon_tuple[2]
                                        if lexicon not in parsing_result_lexicon:
                                            parsing_result_lexicon[lexicon] = {"company_id": lexicon_tuple[0],
                                                                               "word_type": lexicon_tuple[1],
                                                                               "appear_position": [],
                                                                               "count": 0
                                                                              }

                                        appear_found = [m.start() for m in re.finditer(lexicon, cell_value)]

                                        if len(appear_found) > 0:
                                            positions = []

                                            for time in range(len(appear_found)):
                                                positions.append((curr_row+1, curr_cell+1))

                                            origin_position_data = parsing_result_lexicon[lexicon]["appear_position"]
                                            origin_position_data.extend(positions)
                                            parsing_result_lexicon[lexicon]["appear_position"] = origin_position_data

                                            origin_count = parsing_result_lexicon[lexicon]["count"]
                                            origin_count += len(appear_found)
                                            parsing_result_lexicon[lexicon]["count"] = origin_count

                                    #for thulac search
                                    text_cut = self.thu1.cut(cell_value)  # [(tuple)]
                                    filted_cut = []
                                    for ele in text_cut:
                                        if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                            if len(ele[0]) > 1:
                                                filted_cut.append(ele[0])

                                    if len(filted_cut) > 0:
                                        for thulacword in filted_cut:
                                            if thulacword not in parsing_result_thulac:
                                                parsing_result_thulac[thulacword] = {"appear_position": [],
                                                                                     "count": 0
                                                                                    }
                                            thulacword_appear_found = [m.start() for m in re.finditer(thulacword, cell_value)]

                                            if len(thulacword_appear_found) > 0:
                                                thepositions = []
                                                for time in range(len(thulacword_appear_found)):
                                                    thepositions.append((curr_row+1, curr_cell+1))

                                                origin_position_data = parsing_result_thulac[thulacword]["appear_position"]
                                                origin_position_data.extend(thepositions)
                                                parsing_result_thulac[thulacword]["appear_position"] = origin_position_data

                                                origin_count = parsing_result_thulac[thulacword]["count"]
                                                origin_count += len(thulacword_appear_found)
                                                parsing_result_thulac[thulacword]["count"] = origin_count
                                    
                            sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d, '%s', %d, '%s','%s','%s');" % (task_table_name, doc_name, sheet_name, task_id,  task_full_text, task_status, task_start_date, task_end_date, task_code)
                            self.sqlMa.execute(sql_query)

                        #write to table
                        for k, v in parsing_result_lexicon.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"]>0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (lexicon_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)

                        for k, v in parsing_result_thulac.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"] > 0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (thulac_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)




            else: # 使用openpyxl
                filename_without_extension = filename.replace(".xlsx", "").strip()
                self.currFile = filename_without_extension

                excelFile = pyxl.load_workbook(filepath)
                #所有sheet的名字
                sheet_names = excelFile.get_sheet_names()

                #sheet iteration
                for i in range(len(sheet_names)):
                    #sheet
                    work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                    num_rows = work_sheet.max_row
                    num_columns = work_sheet.max_column

                    # 以row跟column數來判斷sheet是否為Empty, 如果empty就不理會
                    if num_columns != 1 and num_rows != 1:
                        sheet_fn = sheet_names[i].strip()
                        self.currSheet = sheet_fn
                        
                        lexicon_table_name = "_".join([filename_without_extension, sheet_fn, "lexicon"])
                        thulac_table_name = "_".join([filename_without_extension, sheet_fn, "thulac"])
                        task_table_name = "_".join([filename_without_extension, "task"])


                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)

                        sql_query = "DELETE FROM %s;" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)


                        #create task table
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (doc_name TEXT, sheet_name TEXT, task_id INT, task_full_text TEXT, task_status INT, task_start_date TEXT, task_end_date TEXT, task_code TEXT);" % (task_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (task_table_name)
                        self.sqlMa.execute(sql_query)


                        
                        parsing_result_thulac = dict()
                        parsing_result_lexicon = dict()
                        
                        for curr_row in range(2, num_rows + 1):  

                            doc_name = filename
                            sheet_name = sheet_fn
                            task_id = curr_row
                            task_full_text = ""
                            task_code = ""
                            task_status = 1
                            task_start_date = ""
                            task_end_date = ""

                            for curr_cell in range(1, num_columns + 1):
                                cell_value = work_sheet.cell(row=curr_row, column=curr_cell).value

                                if cell_value == None:
                                    cell_value = ""

                                task_full_text = " ".join([task_full_text, str(cell_value)]).strip().replace("'", "").replace("\"", "").upper()

                                if curr_cell == 1:
                                    task_code = cell_value

                                if curr_cell == 4:
                                    task_start_date = cell_value
                                
                                if curr_cell == 6:
                                    task_end_date = cell_value




                               
                                cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value)).upper()
                                #for lexicon search
                                for lexicon_tuple in self.all_lexical:
                                    lexicon = lexicon_tuple[2]
                                    if lexicon not in parsing_result_lexicon:
                                        parsing_result_lexicon[lexicon] = {"company_id": lexicon_tuple[0],
                                                                           "word_type"  : lexicon_tuple[1],
                                                                           "appear_position" : [],
                                                                           "count" : 0
                                                                          }
                                    
                                    appear_found = [m.start() for m in re.finditer(lexicon, cell_value)]
                                    if len(appear_found) > 0:
                                        positions = []
                                        for time in range(len(appear_found)):
                                            positions.append((curr_row, curr_cell))
                                        
                                        origin_position_data = parsing_result_lexicon[lexicon]["appear_position"]
                                        origin_position_data.extend(positions)
                                        parsing_result_lexicon[lexicon]["appear_position"] = origin_position_data
                                        
                                        origin_count = parsing_result_lexicon[lexicon]["count"]
                                        origin_count += len(appear_found)
                                        parsing_result_lexicon[lexicon]["count"] = origin_count
                            
                                #for thulac search
                                text_cut = self.thu1.cut(cell_value)  # [(tuple)]
                                filted_cut = []
                                for ele in text_cut:
                                    if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                        if len(ele[0]) > 1:
                                            filted_cut.append(ele[0])
                                
                                if len(filted_cut) > 0:
                                    for thulacword in filted_cut:
                                        if thulacword not in parsing_result_thulac:
                                            parsing_result_thulac[thulacword] = { "appear_position": [],
                                                                                  "count": 0
                                                                                }
                                        thulacword_appear_found = [ m.start() for m in re.finditer(thulacword, cell_value)]
                                        if len(thulacword_appear_found) > 0:
                                            thepositions = []
                                            for time in range(len(thulacword_appear_found)):
                                                thepositions.append((curr_row, curr_cell))

                                            origin_position_data = parsing_result_thulac[thulacword]["appear_position"]
                                            origin_position_data.extend(thepositions)
                                            parsing_result_thulac[thulacword]["appear_position"] = origin_position_data

                                            origin_count = parsing_result_thulac[thulacword]["count"]
                                            origin_count += len(thulacword_appear_found)
                                            parsing_result_thulac[thulacword]["count"] = origin_count
                                
                            
                            sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d, '%s', %d, '%s','%s','%s');" % (task_table_name, doc_name, sheet_name, task_id,  task_full_text , task_status, task_start_date, task_end_date, task_code)
                            self.sqlMa.execute(sql_query)
                        

                        #write to table
                        for k, v in parsing_result_lexicon.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"]>0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (lexicon_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)

                        for k, v in parsing_result_thulac.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"] > 0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (thulac_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)
    """

    """
    def read_file_row_by_row(self, filepath):

        all_sheets = []

        sep_by_name_lexical = dict()

        filename = os.path.basename(filepath)

        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            if self.printFilename:
                print("檔案名稱:", filename)

            if filename.endswith(".xls"):
                # 使用xlrd

                filename_without_extension = filename.replace(
                    ".xls", "").strip()
                # fileDic[filename_without_extension] = []

                excelFile = xl.open_workbook(filepath)
                # 所有sheet的名字
                sheet_names = excelFile.sheet_names()

                # sheet iteration
                for i in range(len(sheet_names)):
                    # sheet
                    work_sheet = excelFile.sheet_by_index(i)

                    # 以row數來判斷sheet是否為Empty
                    if work_sheet.nrows != 0:
                        # print("%s - %s: %d rows" %(filename, sheet_names[i], work_sheet.nrows))
                        # file_create_date = creation_date(src)
                        sheet_fn = sheet_names[i].strip()

                        sheet_contents = ""
                        num_rows = work_sheet.nrows - 1
                        num_cells = work_sheet.ncols - 1
                        curr_row = -1

                        while curr_row < num_rows:
                            curr_row += 1
                            # row = work_sheet.row(curr_row)
                            curr_cell = -1


                            content_in_row = []
                            while curr_cell < num_cells:
                                curr_cell += 1
                                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                                cell_type = work_sheet.cell_type(
                                    curr_row, curr_cell)
                                cell_value = work_sheet.cell_value(
                                    curr_row, curr_cell)

                                try:
                                    if cell_type == 3:
                                        cell_value = int(
                                            cell_value * 24 * 3600)
                                        cell_value = time(
                                            cell_value // 3600, (cell_value % 3600) // 60, cell_value % 60)
                                except:
                                    cell_value = ""

                                if cell_type != 0 and cell_type != 6 and cell_type != 5:
                                    cell_value = re.sub(
                                        "[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                    sheet_contents = " ".join(
                                        [sheet_contents, str(cell_value)]).strip()
                                    
                                    content_in_row.append(cell_value)
                            
                            content_in_row = [x for x in content_in_row if x != '']
                            content_in_string = ' '.join(content_in_row).strip()

                            gotcha = False
                            for name in self.name_lexical:

                                if name in content_in_string:
                                    if name not in sep_by_name_lexical:
                                        sep_by_name_lexical[name] = ''

                                    remainStr = content_in_string.replace(name, '')
                                    if len(remainStr) > 1:
                                        sep_by_name_lexical[name] = ' '.join([sep_by_name_lexical[name], remainStr]).strip()
                                        gotcha = True

                            if not gotcha:
                                if 'unknown' not in sep_by_name_lexical:
                                    sep_by_name_lexical['unknown'] = ''
                                if len(content_in_string)>1:
                                    sep_by_name_lexical['unknown'] = ' '.join([sep_by_name_lexical['unknown'], content_in_string]).strip()

            else:
                # 使用openpyxl
                filename_without_extension = filename.replace(".xlsx", "").strip()

                excelFile = pyxl.load_workbook(filepath)
                #所有sheet的名字
                sheet_names = excelFile.get_sheet_names()

                #sheet iteration
                for i in range(len(sheet_names)):
                    #sheet
                    work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                    num_rows = work_sheet.max_row
                    num_columns = work_sheet.max_column

                    if num_columns != 1 and num_rows != 1:
                        # not empty
                        sheet_contents = ""

                        for curr_row in range(1, num_rows + 1):
                            content_in_row = []
                            for curr_cell in range(1, num_columns + 1):
                                cell_value = work_sheet.cell(
                                    row=curr_row, column=curr_cell).value
                                if cell_value == None:
                                    cell_value = ""

                                cell_value = re.sub(
                                    "[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                sheet_contents = " ".join(
                                    [sheet_contents, str(cell_value)]).strip()
                                content_in_row.append(cell_value)
                            content_in_row = [x for x in content_in_row if x != '']
                            # print(content_in_row)

                            content_in_string = ' '.join(content_in_row).strip()

                            gotcha = False
                            for name in self.name_lexical:
                                
                                if name in content_in_string:
                                    if name not in sep_by_name_lexical:
                                        sep_by_name_lexical[name] = ''

                                    remainStr = content_in_string.replace(name, '')
                                    if len(remainStr) > 1:
                                        sep_by_name_lexical[name] = ' '.join([sep_by_name_lexical[name], remainStr]).strip()
                                        gotcha = True

                            if not gotcha:
                                if 'unknown' not in sep_by_name_lexical:
                                    sep_by_name_lexical['unknown'] = ''
                                if len(content_in_string) > 1:
                                    sep_by_name_lexical['unknown'] = ' '.join([sep_by_name_lexical['unknown'], content_in_string]).strip()

        
        for k, v in sep_by_name_lexical.items():
            # print(k, ':', len(v))
            # print(k)
            text_cut = self.thu1.cut(v)  # [(tuple)]

            filted_str = ""
            for ele in text_cut:
                if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                    if len(ele[0])>1:
                        filted_str = " ".join([filted_str, ele[0]])
            
            text_cut = filted_str.strip()
            freq_result = self.count_frequency(text_cut)
            
            all_sheets.append((k, freq_result))


            # #             # all_sheets.append((sheet_names[i], freq_result))
            # print(freq_result)
            # print("------------------------")


        self.print_out_result(all_sheets)


    def list_all(self, dir_path):

        def print_all_files_sheets(all_files_sheets):
            #print out all:
            for k in all_files_sheets.keys():
                print("檔名:", k)
                print("Sheet名稱:")
                for i in range(len(all_files_sheets[k])):
                    print("\t", all_files_sheets[k][i])
                print("-----------------------------------------------------")


        all_files_sheets = dict()

        curr_path = os.getcwd()

        if os.path.isdir(dir_path):
            os.chdir(dir_path)
            # print(os.getcwd())
            for roots, dirs, files in os.walk(os.getcwd()):
                for filename in files: 
                    #filename:檔案名稱
                    src = os.path.join(roots, filename)
                    if filename.endswith(".xls") or filename.endswith(".xlsx"):
                        if filename not in all_files_sheets:
                            all_files_sheets[filename] = list()

                        if filename.endswith(".xls"):
                            # 使用xlrd
                            excelFile = xl.open_workbook(src)
                            # 所有sheet的名字
                            sheet_names = excelFile.sheet_names()
                            not_empty_sheet_names = []
                            # sheet iteration
                            for i in range(len(sheet_names)):
                                # sheet
                                work_sheet = excelFile.sheet_by_index(i)
                                # 以row數來判斷sheet是否為Empty
                                if work_sheet.nrows != 0:
                                    sheet_name = sheet_names[i].strip()
                                    not_empty_sheet_names.append(sheet_name)
                                    
                            all_files_sheets[filename] = not_empty_sheet_names

                        else:
                            # 使用openpyxl
                            filename_without_extension = filename.replace(".xlsx", "").strip()
                            excelFile = pyxl.load_workbook(src)
                            #所有sheet的名字
                            sheet_names = excelFile.get_sheet_names()
                            not_empty_sheet_names = []

                            #sheet iteration
                            for i in range(len(sheet_names)):
                                #sheet
                                work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                                num_rows = work_sheet.max_row
                                num_columns = work_sheet.max_column

                                if num_columns != 1 and num_rows != 1:
                                    # not empty
                                    sheet_name = sheet_names[i].strip()
                                    not_empty_sheet_names.append(sheet_name)
                                    
                            all_files_sheets[filename] = not_empty_sheet_names
            
            print_all_files_sheets(all_files_sheets)
        
        elif os.path.isfile(dir_path):
            filename = dir_path
            src = os.path.abspath(filename)
            if filename.endswith(".xls") or filename.endswith(".xlsx"):
                if filename not in all_files_sheets:
                    all_files_sheets[filename] = list()

                if filename.endswith(".xls"):
                    # 使用xlrd
                    excelFile = xl.open_workbook(src)
                    # 所有sheet的名字
                    sheet_names = excelFile.sheet_names()
                    not_empty_sheet_names = []
                    # sheet iteration
                    for i in range(len(sheet_names)):
                        # sheet
                        work_sheet = excelFile.sheet_by_index(i)
                        # 以row數來判斷sheet是否為Empty
                        if work_sheet.nrows != 0:
                            sheet_name = sheet_names[i].strip()
                            not_empty_sheet_names.append(sheet_name)

                    all_files_sheets[filename] = not_empty_sheet_names

                else:
                    # 使用openpyxl
                    filename_without_extension = filename.replace(".xlsx", "").strip()
                    excelFile = pyxl.load_workbook(src)
                    #所有sheet的名字
                    sheet_names = excelFile.get_sheet_names()
                    not_empty_sheet_names = []

                    #sheet iteration
                    for i in range(len(sheet_names)):
                        #sheet
                        work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                        num_rows = work_sheet.max_row
                        num_columns = work_sheet.max_column

                        if num_columns != 1 and num_rows != 1:
                            # not empty
                            sheet_name = sheet_names[i].strip()
                            not_empty_sheet_names.append(sheet_name)

                    all_files_sheets[filename] = not_empty_sheet_names
        
            print_all_files_sheets(all_files_sheets)
        else:
            print("sorry, " + str(dir_path) + " not a directory path")

        os.chdir(curr_path)

    def go_through_directory(self, dir_path):
        os.chdir(dir_path)
        # print(os.getcwd())
        for roots, dirs, files in os.walk(os.getcwd()):
            for filename in files:
                src = os.path.join(roots, filename)
                # print(src)
                self.print_out_result(self.read_single_file(src))

    """