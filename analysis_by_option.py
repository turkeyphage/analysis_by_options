#!/Library/Frameworks/Python.framework/Versions/3.5/bin/python3
import os
import os.path
import sys
import xlrd as xl
import openpyxl as pyxl
import errno
import re
import time
import platform
import datetime
from datetime import time
import thulac
import optparse
import collections
from sqlite_manager import SQLiteManager

class FileAnalyzer:
    def __init__(self):
        self.thu1 = thulac.thulac(filt=True)
        self.printFilename = True
        self.sqlMa = SQLiteManager()
        # self.name_lexical = []
        self.all_lexical = []
        # print(self.all_lexical)
        for ele in self.sqlMa.execute("SELECT * FROM lexical;"):
            
            # self.all_lexical.append(ele)
            for ele_in_tuple in ele:
                self.all_lexical.append(ele_in_tuple)

        # print(self.all_lexical)

        # self.sqlMa.execute("SELECT * FROM lexical;")
      
    def creation_date(self, path_to_file):

        if platform.system() == 'Windows':
            return os.path.getctime(path_to_file)
        else:
            stat = os.stat(path_to_file)
            try:
                full_datetime = datetime.datetime.fromtimestamp(stat.st_birthtime)
                year = full_datetime.year
                month = full_datetime.month
                day = full_datetime.day
                return "-".join([str(year), str(month), str(day)])
            except AttributeError:
                # We're probably on Linux. No easy way to get creation dates here,
                # so we'll settle for when its content was last modified.
                full_datetime = datetime.datetime.fromtimestamp(stat.st_mtime)
                year = full_datetime.year
                month = full_datetime.month
                day = full_datetime.day
                return "-".join([str(year), str(month), str(day)])


    def count_frequency(self, full_text):

        split_text = full_text.split(" ")

        feq_count = dict()

        for each in split_text:
            if each not in feq_count:
                feq_count[each] = 0

            feq_count[each] = feq_count[each] + 1

        feq_combined = dict()
        for w in sorted(feq_count, key=feq_count.get, reverse=True):
            if feq_count[w] not in feq_combined:
                feq_combined[feq_count[w]] = ""
            feq_combined[feq_count[w]] = ",".join([feq_combined[feq_count[w]], w]).lstrip(',')

        return feq_combined


    #列印結果
    def print_out_result(self, allsheets):
        
        for i in range(len(allsheets)):
            sheet_name = allsheets[i][0]
            od = collections.OrderedDict(sorted(allsheets[i][1].items()))
            # print("sheet_name: " + sheet_name)
            print("<" + sheet_name + ">")
            print("字頻統計: ")
            for k, v in od.items():
                if k > 5:
                    print(str(k) + "次: " + v)
            print("-----------------------------------------------------")



    #單一檔案讀取
    def read_single_file(self, filepath):
        all_sheets = []
        # thu1 = thulac.thulac(seg_only=True, filt=True)
        
        filename = os.path.basename(filepath)

        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            if self.printFilename:
                print("檔案名稱:", filename)
            # thu1 = thulac.thulac(filt=True)

            if filename.endswith(".xls"):
                # 使用xlrd
                
                filename_without_extension = filename.replace(".xls", "").strip()
                # fileDic[filename_without_extension] = []

                excelFile = xl.open_workbook(filepath)
                # 所有sheet的名字
                sheet_names = excelFile.sheet_names()

                # sheet iteration
                for i in range(len(sheet_names)):
                    # sheet
                    work_sheet = excelFile.sheet_by_index(i)

                    # 以row數來判斷sheet是否為Empty
                    if work_sheet.nrows != 0:
                        # print("%s - %s: %d rows" %(filename, sheet_names[i], work_sheet.nrows))
                        # file_create_date = creation_date(src)
                        sheet_fn = sheet_names[i].strip()

                        sheet_contents = ""
                        num_rows = work_sheet.nrows - 1
                        num_cells = work_sheet.ncols - 1
                        curr_row = -1

                        while curr_row < num_rows:
                            curr_row += 1
                            # row = work_sheet.row(curr_row)
                            curr_cell = -1

                            while curr_cell < num_cells:
                                curr_cell += 1
                                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                                cell_type = work_sheet.cell_type(curr_row, curr_cell)
                                cell_value = work_sheet.cell_value(curr_row, curr_cell)

                                try:
                                    if cell_type == 3:
                                        cell_value = int(cell_value * 24 * 3600)
                                        cell_value = time(cell_value // 3600, (cell_value % 3600) // 60, cell_value % 60)
                                except:
                                    cell_value = ""

                                if cell_type != 0 and cell_type != 6 and cell_type != 5:
                                    cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                    sheet_contents = " ".join(
                                        [sheet_contents, str(cell_value)]).strip()

                        # text_cut = thu1.cut(sheet_contents, text=True).strip()

                        text_cut = self.thu1.cut(sheet_contents) #[(tuple)]


                        filted_str = ""
                        for ele in text_cut:
                            if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                filted_str = " ".join([filted_str, ele[0]])
                        text_cut = filted_str.strip()
                        freq_result = self.count_frequency(text_cut)
                        
                        all_sheets.append((sheet_names[i], freq_result))


            else:
                # 使用openpyxl
                filename_without_extension = filename.replace(".xlsx", "").strip()

                excelFile = pyxl.load_workbook(filepath)
                #所有sheet的名字
                sheet_names = excelFile.get_sheet_names()

                #sheet iteration
                for i in range(len(sheet_names)):
                    #sheet
                    work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                    num_rows = work_sheet.max_row
                    num_columns = work_sheet.max_column

                    if num_columns != 1 and num_rows != 1:
                        # not empty
                        sheet_contents = ""

                        for curr_row in range(1, num_rows + 1):
                            for curr_cell in range(1, num_columns + 1):
                                cell_value = work_sheet.cell(
                                    row=curr_row, column=curr_cell).value
                                if cell_value == None:
                                    cell_value = ""

                                cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                sheet_contents = " ".join(
                                    [sheet_contents, str(cell_value)]).strip()
                                
                        # text_cut = thu1.cut(sheet_contents, text=True).strip()
                        text_cut = self.thu1.cut(sheet_contents)
                        filted_str = ""
                        for ele in text_cut:
                            if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                filted_str = " ".join([filted_str, ele[0]])
                        text_cut = filted_str.strip()
                        freq_result = self.count_frequency(text_cut)

                        all_sheets.append((sheet_names[i], freq_result))

        return all_sheets  # [(2 items tuple)]



    def parse_row_by_row(self, filepath):
        
        filename = os.path.basename(filepath)

        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            if self.printFilename:
                print("檔案名稱:", filename)

            if filename.endswith(".xls"): # 使用xlrd
                filename_without_extension = filename.replace(".xls", "").strip()
                excelFile = xl.open_workbook(filepath)
                # 所有sheet的名字
                sheet_names = excelFile.sheet_names()

                # sheet iteration
                for i in range(len(sheet_names)):
                    # sheet
                    
                    work_sheet = excelFile.sheet_by_index(i)

                    # 以row數來判斷sheet是否為Empty, 如果empty就不理會
                    if work_sheet.nrows != 0:

                        #sheet名
                        sheet_fn = sheet_names[i].strip()

                        num_rows = work_sheet.nrows
                        num_columns = work_sheet.ncols
                        
                        lexicon_table_name = "_".join([filename_without_extension, sheet_fn, "lexicon"])
                        thulac_table_name = "_".join([filename_without_extension, sheet_fn, "thulac"])
                        task_table_name = "_".join([filename_without_extension, "task"])

                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)

                        sql_query = "DELETE FROM %s;" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)


                        #create task table
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (doc_name TEXT, sheet_name TEXT, task_id INT, task_full_text TEXT, task_status INT, task_start_date TEXT, task_end_date TEXT, task_code TEXT);" % (task_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (task_table_name)
                        self.sqlMa.execute(sql_query)

                        parsing_result_thulac = dict()
                        parsing_result_lexicon = dict()

                        for curr_row in range(1, num_rows):

                            doc_name = filename
                            sheet_name = sheet_fn
                            task_id = curr_row+1
                            task_full_text = ""
                            task_code = ""
                            task_status = 1
                            task_start_date = ""
                            task_end_date = ""


                            for curr_cell in range(0, num_columns):
                                
                                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                                cell_type = work_sheet.cell_type(curr_row, curr_cell)
                                cell_value = work_sheet.cell_value(curr_row, curr_cell)

                                try:
                                    if cell_type == 3:
                                        cell_value = int(cell_value * 24 * 3600)
                                        cell_value = time(cell_value // 3600, (cell_value % 3600) // 60, cell_value % 60)
                                except:
                                    cell_value = ""

                                task_full_text = " ".join([task_full_text, str(cell_value)]).strip().replace("'", "").replace("\"", "").upper()

                                if curr_cell == 0:
                                    task_code = cell_value

                                if curr_cell == 3:
                                    task_start_date = cell_value
                                
                                if curr_cell == 5:
                                    task_end_date = cell_value




                                if cell_type != 0 and cell_type != 6 and cell_type != 5:
                                    cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value)).upper()
                                    
                                    #for lecial search
                                    for lexicon_tuple in self.all_lexical:
                                        lexicon = lexicon_tuple[2]
                                        if lexicon not in parsing_result_lexicon:
                                            parsing_result_lexicon[lexicon] = {"company_id": lexicon_tuple[0],
                                                                               "word_type": lexicon_tuple[1],
                                                                               "appear_position": [],
                                                                               "count": 0
                                                                              }

                                        appear_found = [m.start() for m in re.finditer(lexicon, cell_value)]

                                        if len(appear_found) > 0:
                                            positions = []

                                            for time in range(len(appear_found)):
                                                positions.append((curr_row+1, curr_cell+1))

                                            origin_position_data = parsing_result_lexicon[lexicon]["appear_position"]
                                            origin_position_data.extend(positions)
                                            parsing_result_lexicon[lexicon]["appear_position"] = origin_position_data

                                            origin_count = parsing_result_lexicon[lexicon]["count"]
                                            origin_count += len(appear_found)
                                            parsing_result_lexicon[lexicon]["count"] = origin_count

                                    #for thulac search
                                    text_cut = self.thu1.cut(cell_value)  # [(tuple)]
                                    filted_cut = []
                                    for ele in text_cut:
                                        if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                            if len(ele[0]) > 1:
                                                filted_cut.append(ele[0])

                                    if len(filted_cut) > 0:
                                        for thulacword in filted_cut:
                                            if thulacword not in parsing_result_thulac:
                                                parsing_result_thulac[thulacword] = {"appear_position": [],
                                                                                     "count": 0
                                                                                    }
                                            thulacword_appear_found = [m.start() for m in re.finditer(thulacword, cell_value)]

                                            if len(thulacword_appear_found) > 0:
                                                thepositions = []
                                                for time in range(len(thulacword_appear_found)):
                                                    thepositions.append((curr_row+1, curr_cell+1))

                                                origin_position_data = parsing_result_thulac[thulacword]["appear_position"]
                                                origin_position_data.extend(thepositions)
                                                parsing_result_thulac[thulacword]["appear_position"] = origin_position_data

                                                origin_count = parsing_result_thulac[thulacword]["count"]
                                                origin_count += len(thulacword_appear_found)
                                                parsing_result_thulac[thulacword]["count"] = origin_count
                                    
                            sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d, '%s', %d, '%s','%s','%s');" % (task_table_name, doc_name, sheet_name, task_id,  task_full_text, task_status, task_start_date, task_end_date, task_code)
                            self.sqlMa.execute(sql_query)

                        #write to table
                        for k, v in parsing_result_lexicon.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"]>0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (lexicon_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)

                        for k, v in parsing_result_thulac.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"] > 0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (thulac_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)




            else: # 使用openpyxl
                filename_without_extension = filename.replace(".xlsx", "").strip()
                excelFile = pyxl.load_workbook(filepath)
                #所有sheet的名字
                sheet_names = excelFile.get_sheet_names()

                #sheet iteration
                for i in range(len(sheet_names)):
                    #sheet
                    work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                    num_rows = work_sheet.max_row
                    num_columns = work_sheet.max_column

                    # 以row跟column數來判斷sheet是否為Empty, 如果empty就不理會
                    if num_columns != 1 and num_rows != 1:
                        sheet_fn = sheet_names[i].strip()
                        
                        lexicon_table_name = "_".join([filename_without_extension, sheet_fn, "lexicon"])
                        thulac_table_name = "_".join([filename_without_extension, sheet_fn, "thulac"])
                        task_table_name = "_".join([filename_without_extension, "task"])


                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (keyword TEXT, position TEXT, count INT);" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)

                        sql_query = "DELETE FROM %s;" % (lexicon_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (thulac_table_name)
                        self.sqlMa.execute(sql_query)


                        #create task table
                        sql_query = "CREATE TABLE IF NOT EXISTS %s (doc_name TEXT, sheet_name TEXT, task_id INT, task_full_text TEXT, task_status INT, task_start_date TEXT, task_end_date TEXT, task_code TEXT);" % (task_table_name)
                        self.sqlMa.execute(sql_query)
                        sql_query = "DELETE FROM %s;" % (task_table_name)
                        self.sqlMa.execute(sql_query)


                        
                        parsing_result_thulac = dict()
                        parsing_result_lexicon = dict()
                        
                        for curr_row in range(2, num_rows + 1):  

                            doc_name = filename
                            sheet_name = sheet_fn
                            task_id = curr_row
                            task_full_text = ""
                            task_code = ""
                            task_status = 1
                            task_start_date = ""
                            task_end_date = ""

                            for curr_cell in range(1, num_columns + 1):
                                cell_value = work_sheet.cell(row=curr_row, column=curr_cell).value

                                if cell_value == None:
                                    cell_value = ""

                                task_full_text = " ".join([task_full_text, str(cell_value)]).strip().replace("'", "").replace("\"", "").upper()

                                if curr_cell == 1:
                                    task_code = cell_value

                                if curr_cell == 4:
                                    task_start_date = cell_value
                                
                                if curr_cell == 6:
                                    task_end_date = cell_value




                               
                                cell_value = re.sub("[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value)).upper()
                                #for lexicon search
                                for lexicon_tuple in self.all_lexical:
                                    lexicon = lexicon_tuple[2]
                                    if lexicon not in parsing_result_lexicon:
                                        parsing_result_lexicon[lexicon] = {"company_id": lexicon_tuple[0],
                                                                           "word_type"  : lexicon_tuple[1],
                                                                           "appear_position" : [],
                                                                           "count" : 0
                                                                          }
                                    
                                    appear_found = [m.start() for m in re.finditer(lexicon, cell_value)]
                                    if len(appear_found) > 0:
                                        positions = []
                                        for time in range(len(appear_found)):
                                            positions.append((curr_row, curr_cell))
                                        
                                        origin_position_data = parsing_result_lexicon[lexicon]["appear_position"]
                                        origin_position_data.extend(positions)
                                        parsing_result_lexicon[lexicon]["appear_position"] = origin_position_data
                                        
                                        origin_count = parsing_result_lexicon[lexicon]["count"]
                                        origin_count += len(appear_found)
                                        parsing_result_lexicon[lexicon]["count"] = origin_count
                            
                                #for thulac search
                                text_cut = self.thu1.cut(cell_value)  # [(tuple)]
                                filted_cut = []
                                for ele in text_cut:
                                    if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                                        if len(ele[0]) > 1:
                                            filted_cut.append(ele[0])
                                
                                if len(filted_cut) > 0:
                                    for thulacword in filted_cut:
                                        if thulacword not in parsing_result_thulac:
                                            parsing_result_thulac[thulacword] = { "appear_position": [],
                                                                                  "count": 0
                                                                                }
                                        thulacword_appear_found = [ m.start() for m in re.finditer(thulacword, cell_value)]
                                        if len(thulacword_appear_found) > 0:
                                            thepositions = []
                                            for time in range(len(thulacword_appear_found)):
                                                thepositions.append((curr_row, curr_cell))

                                            origin_position_data = parsing_result_thulac[thulacword]["appear_position"]
                                            origin_position_data.extend(thepositions)
                                            parsing_result_thulac[thulacword]["appear_position"] = origin_position_data

                                            origin_count = parsing_result_thulac[thulacword]["count"]
                                            origin_count += len(thulacword_appear_found)
                                            parsing_result_thulac[thulacword]["count"] = origin_count
                                
                            
                            sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d, '%s', %d, '%s','%s','%s');" % (task_table_name, doc_name, sheet_name, task_id,  task_full_text , task_status, task_start_date, task_end_date, task_code)
                            self.sqlMa.execute(sql_query)
                        

                        #write to table
                        for k, v in parsing_result_lexicon.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"]>0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (lexicon_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)

                        for k, v in parsing_result_thulac.items():
                            #k = keyword
                            #value = dictionary
                            keyword = k
                            pos_str = ""
                            if v["count"] > 0:
                                for pos_tuple in v["appear_position"]:
                                    # pos = "[" + str(pos_tuple[0]) + "," + str(pos_tuple[-1]) + "]"
                                    pos = str(pos_tuple[0])
                                    pos_str = ','.join([pos_str, pos]).strip(',')
                                count = v["count"]
                                sql_query = "INSERT INTO %s VALUES ('%s', '%s', %d);" % (thulac_table_name, k, pos_str, count)
                                self.sqlMa.execute(sql_query)

                        
                        



    def read_file_row_by_row(self, filepath):

        all_sheets = []

        sep_by_name_lexical = dict()

        filename = os.path.basename(filepath)

        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            if self.printFilename:
                print("檔案名稱:", filename)

            if filename.endswith(".xls"):
                # 使用xlrd

                filename_without_extension = filename.replace(
                    ".xls", "").strip()
                # fileDic[filename_without_extension] = []

                excelFile = xl.open_workbook(filepath)
                # 所有sheet的名字
                sheet_names = excelFile.sheet_names()

                # sheet iteration
                for i in range(len(sheet_names)):
                    # sheet
                    work_sheet = excelFile.sheet_by_index(i)

                    # 以row數來判斷sheet是否為Empty
                    if work_sheet.nrows != 0:
                        # print("%s - %s: %d rows" %(filename, sheet_names[i], work_sheet.nrows))
                        # file_create_date = creation_date(src)
                        sheet_fn = sheet_names[i].strip()

                        sheet_contents = ""
                        num_rows = work_sheet.nrows - 1
                        num_cells = work_sheet.ncols - 1
                        curr_row = -1

                        while curr_row < num_rows:
                            curr_row += 1
                            # row = work_sheet.row(curr_row)
                            curr_cell = -1


                            content_in_row = []
                            while curr_cell < num_cells:
                                curr_cell += 1
                                # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
                                cell_type = work_sheet.cell_type(
                                    curr_row, curr_cell)
                                cell_value = work_sheet.cell_value(
                                    curr_row, curr_cell)

                                try:
                                    if cell_type == 3:
                                        cell_value = int(
                                            cell_value * 24 * 3600)
                                        cell_value = time(
                                            cell_value // 3600, (cell_value % 3600) // 60, cell_value % 60)
                                except:
                                    cell_value = ""

                                if cell_type != 0 and cell_type != 6 and cell_type != 5:
                                    cell_value = re.sub(
                                        "[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                    sheet_contents = " ".join(
                                        [sheet_contents, str(cell_value)]).strip()
                                    
                                    content_in_row.append(cell_value)
                            
                            content_in_row = [x for x in content_in_row if x != '']
                            content_in_string = ' '.join(content_in_row).strip()

                            gotcha = False
                            for name in self.name_lexical:

                                if name in content_in_string:
                                    if name not in sep_by_name_lexical:
                                        sep_by_name_lexical[name] = ''

                                    remainStr = content_in_string.replace(name, '')
                                    if len(remainStr) > 1:
                                        sep_by_name_lexical[name] = ' '.join([sep_by_name_lexical[name], remainStr]).strip()
                                        gotcha = True

                            if not gotcha:
                                if 'unknown' not in sep_by_name_lexical:
                                    sep_by_name_lexical['unknown'] = ''
                                if len(content_in_string)>1:
                                    sep_by_name_lexical['unknown'] = ' '.join([sep_by_name_lexical['unknown'], content_in_string]).strip()

            else:
                # 使用openpyxl
                filename_without_extension = filename.replace(".xlsx", "").strip()

                excelFile = pyxl.load_workbook(filepath)
                #所有sheet的名字
                sheet_names = excelFile.get_sheet_names()

                #sheet iteration
                for i in range(len(sheet_names)):
                    #sheet
                    work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                    num_rows = work_sheet.max_row
                    num_columns = work_sheet.max_column

                    if num_columns != 1 and num_rows != 1:
                        # not empty
                        sheet_contents = ""

                        for curr_row in range(1, num_rows + 1):
                            content_in_row = []
                            for curr_cell in range(1, num_columns + 1):
                                cell_value = work_sheet.cell(
                                    row=curr_row, column=curr_cell).value
                                if cell_value == None:
                                    cell_value = ""

                                cell_value = re.sub(
                                    "[\s\d，,<>;：（）().；\-:、/.。《》…~\"\\\%$&#*@~`？]+", "", str(cell_value))
                                sheet_contents = " ".join(
                                    [sheet_contents, str(cell_value)]).strip()
                                content_in_row.append(cell_value)
                            content_in_row = [x for x in content_in_row if x != '']
                            # print(content_in_row)

                            content_in_string = ' '.join(content_in_row).strip()

                            gotcha = False
                            for name in self.name_lexical:
                                
                                if name in content_in_string:
                                    if name not in sep_by_name_lexical:
                                        sep_by_name_lexical[name] = ''

                                    remainStr = content_in_string.replace(name, '')
                                    if len(remainStr) > 1:
                                        sep_by_name_lexical[name] = ' '.join([sep_by_name_lexical[name], remainStr]).strip()
                                        gotcha = True

                            if not gotcha:
                                if 'unknown' not in sep_by_name_lexical:
                                    sep_by_name_lexical['unknown'] = ''
                                if len(content_in_string) > 1:
                                    sep_by_name_lexical['unknown'] = ' '.join([sep_by_name_lexical['unknown'], content_in_string]).strip()

        
        for k, v in sep_by_name_lexical.items():
            # print(k, ':', len(v))
            # print(k)
            text_cut = self.thu1.cut(v)  # [(tuple)]

            filted_str = ""
            for ele in text_cut:
                if ele[-1] != 'u' and ele[-1] != 'y' and ele[-1] != 'r':
                    if len(ele[0])>1:
                        filted_str = " ".join([filted_str, ele[0]])
            
            text_cut = filted_str.strip()
            freq_result = self.count_frequency(text_cut)
            
            all_sheets.append((k, freq_result))


            # #             # all_sheets.append((sheet_names[i], freq_result))
            # print(freq_result)
            # print("------------------------")


        self.print_out_result(all_sheets)


    def list_all(self, dir_path):

        def print_all_files_sheets(all_files_sheets):
            #print out all:
            for k in all_files_sheets.keys():
                print("檔名:", k)
                print("Sheet名稱:")
                for i in range(len(all_files_sheets[k])):
                    print("\t", all_files_sheets[k][i])
                print("-----------------------------------------------------")


        all_files_sheets = dict()

        curr_path = os.getcwd()

        if os.path.isdir(dir_path):
            os.chdir(dir_path)
            # print(os.getcwd())
            for roots, dirs, files in os.walk(os.getcwd()):
                for filename in files: 
                    #filename:檔案名稱
                    src = os.path.join(roots, filename)
                    if filename.endswith(".xls") or filename.endswith(".xlsx"):
                        if filename not in all_files_sheets:
                            all_files_sheets[filename] = list()

                        if filename.endswith(".xls"):
                            # 使用xlrd
                            excelFile = xl.open_workbook(src)
                            # 所有sheet的名字
                            sheet_names = excelFile.sheet_names()
                            not_empty_sheet_names = []
                            # sheet iteration
                            for i in range(len(sheet_names)):
                                # sheet
                                work_sheet = excelFile.sheet_by_index(i)
                                # 以row數來判斷sheet是否為Empty
                                if work_sheet.nrows != 0:
                                    sheet_name = sheet_names[i].strip()
                                    not_empty_sheet_names.append(sheet_name)
                                    
                            all_files_sheets[filename] = not_empty_sheet_names

                        else:
                            # 使用openpyxl
                            filename_without_extension = filename.replace(".xlsx", "").strip()
                            excelFile = pyxl.load_workbook(src)
                            #所有sheet的名字
                            sheet_names = excelFile.get_sheet_names()
                            not_empty_sheet_names = []

                            #sheet iteration
                            for i in range(len(sheet_names)):
                                #sheet
                                work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                                num_rows = work_sheet.max_row
                                num_columns = work_sheet.max_column

                                if num_columns != 1 and num_rows != 1:
                                    # not empty
                                    sheet_name = sheet_names[i].strip()
                                    not_empty_sheet_names.append(sheet_name)
                                    
                            all_files_sheets[filename] = not_empty_sheet_names
            
            print_all_files_sheets(all_files_sheets)
        
        elif os.path.isfile(dir_path):
            filename = dir_path
            src = os.path.abspath(filename)
            if filename.endswith(".xls") or filename.endswith(".xlsx"):
                if filename not in all_files_sheets:
                    all_files_sheets[filename] = list()

                if filename.endswith(".xls"):
                    # 使用xlrd
                    excelFile = xl.open_workbook(src)
                    # 所有sheet的名字
                    sheet_names = excelFile.sheet_names()
                    not_empty_sheet_names = []
                    # sheet iteration
                    for i in range(len(sheet_names)):
                        # sheet
                        work_sheet = excelFile.sheet_by_index(i)
                        # 以row數來判斷sheet是否為Empty
                        if work_sheet.nrows != 0:
                            sheet_name = sheet_names[i].strip()
                            not_empty_sheet_names.append(sheet_name)

                    all_files_sheets[filename] = not_empty_sheet_names

                else:
                    # 使用openpyxl
                    filename_without_extension = filename.replace(".xlsx", "").strip()
                    excelFile = pyxl.load_workbook(src)
                    #所有sheet的名字
                    sheet_names = excelFile.get_sheet_names()
                    not_empty_sheet_names = []

                    #sheet iteration
                    for i in range(len(sheet_names)):
                        #sheet
                        work_sheet = excelFile.get_sheet_by_name(sheet_names[i])

                        num_rows = work_sheet.max_row
                        num_columns = work_sheet.max_column

                        if num_columns != 1 and num_rows != 1:
                            # not empty
                            sheet_name = sheet_names[i].strip()
                            not_empty_sheet_names.append(sheet_name)

                    all_files_sheets[filename] = not_empty_sheet_names
        
            print_all_files_sheets(all_files_sheets)
        else:
            print("sorry, " + str(dir_path) + " not a directory path")

        os.chdir(curr_path)

    def go_through_directory(self, dir_path):
        os.chdir(dir_path)
        # print(os.getcwd())
        for roots, dirs, files in os.walk(os.getcwd()):
            for filename in files:
                src = os.path.join(roots, filename)
                # print(src)
                self.print_out_result(self.read_single_file(src))



    


def main():

    analyzer = FileAnalyzer()
    # sqlitemanager = SQLiteManager()
    # print(sqlitemanager.execute("SELECT * FROM lexical;"))

    parser = optparse.OptionParser(usage = "%prog [options] [file_name or folder_name]\n")

    parser.add_option("-f", "--file",
                      action="store_true",
                      default=False,
                      dest="single_file",
                      help="分析單獨excel檔")

    parser.add_option("-d", "--directory",
                      action="store_true",
                      default=False,
                      dest="all_in_directory",
                      help="分析資料夾裡所有excel檔")
                      
    parser.add_option("-l", "--list",
                      action="store_true",
                      default=False,
                      dest="list_all",
                      help="列出所有excel及sheet名")

    parser.add_option("-c", "--combinedFiles",
                      action="store_true",
                      default=False,
                      dest="combined_Files",
                      help="綜合分析指定excel檔")


    (options, args) = parser.parse_args()

    
        
    if options.single_file:
        if len(args) < 1:
            print("not enough args")
            parser.print_help()
            return -1
        else:
            for i in range(len(args)):
                # analyzer.print_out_result(analyzer.read_single_file(args[i]))
                analyzer.parse_row_by_row(args[i])
            print("分析完畢，已將分析結果存入lexicon.db中")
                




    elif options.all_in_directory:
        if len(args) < 1:
            print("not enough args")
            parser.print_help()
            return -1
        else:
            for i in range(len(args)):
               analyzer.go_through_directory(args[i])
                





    elif options.combined_Files:
        analyzer.printFilename = False
        all_data = dict()
        for i in range(len(args)):
            single_file_data = analyzer.read_single_file(args[i])
            # print(single_file_data)
            
            for ele in single_file_data:
                sn = ele[0]
                tokens_in_sheet = ""  # sheet為單位
                for theKey in ele[1].keys():
                    loopTime = int(theKey)
                    wannaStr = ""
                    for ti in range(loopTime):
                        wannaStr = ",".join([wannaStr, ele[1][theKey]])
                    wannaStr = wannaStr.lstrip(",").strip() #分次字數統計
                    
                    tokens_in_sheet = ",".join([tokens_in_sheet, wannaStr])
                tokens_in_sheet = tokens_in_sheet.lstrip(",").strip()
                    
                # print(sn, ":", tokens_in_sheet)
                
                if sn not in all_data:
                    all_data[sn] = tokens_in_sheet
                else:
                    ori_sn_data = all_data[sn]
                    all_data[sn] = ",".join([ori_sn_data, tokens_in_sheet])

        for k, v in all_data.items():
            cleanV = v.replace(","," ")
            count_result = analyzer.count_frequency(cleanV)
            print(k)
            # print(count_result)

            od = collections.OrderedDict(sorted(count_result.items()))
            print("字頻統計: ")
            for od_k, od_v in od.items():
                print(str(od_k) + "次: " + od_v)
            print("-----------------------------------------------------")


    
    elif options.list_all:
        if len(args) < 1:
            print("not enough args")
            parser.print_help()
            return -1
        else:
            for i in range(len(args)):
                analyzer.list_all(args[i])

    else:
        parser.print_help()

        
if __name__ == "__main__":
    main()
